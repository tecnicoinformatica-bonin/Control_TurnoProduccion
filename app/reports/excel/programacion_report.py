from io import BytesIO
from datetime import datetime

from flask_login import login_required, current_user
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image

from copy import copy

@login_required
def generar_reporte_programacion(programacion, detalles):
    ruta_plantilla = "app/templates/excel/programacion.xlsm"

    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws = wb["Programación"]

    ws["G4"] = programacion["fecha"]
    ws["G5"] = programacion["nombreDepartamento"]
    ws["G6"] = programacion["nombre_elaborado_por"]
    ws["P6"] = programacion["esFeriado"]
    
    # =========================================================
    # DETALLE EMPLEADOS
    # =========================================================

    fila_inicio = 11
    fila_contador = fila_inicio
    fila_plantilla = 11

    contador = 1

    if len(detalles) > 1:
        ws.insert_rows(fila_inicio + 1, len(detalles) - 1)

    for fila in range(fila_inicio + 1, fila_inicio + len(detalles)):
        copiar_estilo_fila(ws, fila_plantilla, fila)

    for detalle in detalles:

        ws[f"A{fila_contador}"] = contador

        ws[f"B{fila_contador}"] = detalle["badgeNumber"]
        
        ws[f"C{fila_contador}"] = detalle["idEmpleado"]

        ws[f"D{fila_contador}"] =  detalle["nombreEmpleado"] #f"{detalle["firstName"]} {detalle["secondName"] or ""} {detalle["lastName"]} {detalle["lastName2"] or ""}"

        ws[f"E{fila_contador}"] = detalle["nombreCentro"]

        ws[f"F{fila_contador}"] = detalle["nombreLinea"]

        ws[f"G{fila_contador}"] = detalle["nombreProceso"]

        ws[f"H{fila_contador}"] = detalle["hora_inicio"]

        ws[f"I{fila_contador}"] = detalle["hora_fin"]
        
        ws[f"J{fila_contador}"] = "SÍ" if detalle["aplica_almuerzo"] else "NO"
        
        ws[f"K{fila_contador}"] = "SÍ" if detalle["aplica_cena"] else "NO"
        
        ws[f"L{fila_contador}"] = "SÍ" if detalle["cena_con_costo"] else "NO"
        
        ws[f"M{fila_contador}"] = "SÍ" if detalle["aplica_transporte"] == 1  else "NO"
        
        ws[f"N{fila_contador}"] = detalle["observacion_transporte"] if detalle["observacion_transporte"] else "----"

        fila_contador += 1
        contador += 1

    ws[f"J{fila_contador}"] = f'=COUNTIF(J{fila_inicio}:J{fila_contador - 1},"SÍ")'
    ws[f"K{fila_contador}"] = f'=COUNTIF(K{fila_inicio}:K{fila_contador - 1},"SÍ")'
    ws[f"L{fila_contador}"] = f'=COUNTIF(L{fila_inicio}:L{fila_contador - 1},"SÍ")'

    ws["E6"] = f"=K{fila_contador}"
    ws["E7"] = f"=L{fila_contador}"
    ws["E8"] = f"=J{fila_contador}"

    # =========================================================
    # ANCHO COLUMNAS
    # =========================================================

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 15
    ws.column_dimensions["N"].width = 45
    ws.column_dimensions["O"].width = 20
    ws.column_dimensions["P"].width = 20

    # =========================================================
    # PIE DE REPORTE
    # =========================================================

    fila_footer = fila_contador + 2

    ws[f"N{fila_footer}"] = "Generado el:"

    ws[f"O{fila_footer}"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws[f"N{fila_footer + 1}"] = "Por:"

    ws[f"O{fila_footer + 1}"] = current_user.nombre

    # =========================================================
    # EXPORTAR EN MEMORIA
    # =========================================================
    
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return archivo

def copiar_estilo_fila(ws, fila_origen, fila_destino):

    for col in range(1, ws.max_column + 1):

        celda_origen = ws.cell(row=fila_origen, column=col)
        celda_destino = ws.cell(row=fila_destino, column=col)

        # Copiar estilo
        celda_destino.font = copy(celda_origen.font)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)

        # Copiar altura fila
        ws.row_dimensions[fila_destino].height = ws.row_dimensions[fila_origen].height