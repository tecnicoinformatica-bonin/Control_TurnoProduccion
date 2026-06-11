from io import BytesIO
from datetime import datetime, timedelta

from flask_login import login_required, current_user
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from copy import copy

import pytz

def copiar_estilo_fila(ws, fila_origen, fila_destino):

    for col in range(1, ws.max_column + 1):

        celda_origen = ws.cell(row=fila_origen, column=col)
        celda_destino = ws.cell(row=fila_destino, column=col)

        # Copiar estilo
        celda_destino.font = copy(celda_origen.font)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)

        # Copiar altura fila
        ws.row_dimensions[fila_destino].height = ws.row_dimensions[fila_origen].height

def copiar_estilo_columna(ws, col_origen, col_destino, fila_inicio, fila_fin):

    for fila in range(fila_inicio, fila_fin + 1):

        origen = ws.cell(fila, col_origen)
        destino = ws.cell(fila, col_destino)

        destino.font = copy(origen.font)
        destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format
        destino.protection = copy(origen.protection)

def obtener_diferencia_fechas(from_date, to_date):
        fecha_inicio = datetime.strptime(from_date, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(to_date, "%Y-%m-%d").date()

        fechas = []

        dias = (fecha_fin - fecha_inicio).days + 1

        fechas = [
            fecha_inicio + timedelta(days=i)
            for i in range(dias)
        ]
        
        return fechas

@login_required
def generar_reporte_horas_extra_autorizadas(encabezado_detalles, detalles):
    ruta_plantilla = "app/templates/excel/horas_extra_autorizadas.xlsm"

    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws = wb["Resumen"]

    ws["D2"] = encabezado_detalles["nombreDepartamento"]
    ws["D3"] = encabezado_detalles["from_date"]
    ws["D4"] = encabezado_detalles["to_date"]
    
    
    # =========================================================
    # DETALLE EMPLEADOS
    # =========================================================

    fila_inicio = 10
    fila_contador = fila_inicio
    fila_plantilla = 10

    contador = 1

    if len(detalles) > 1:
        ws.insert_rows(fila_inicio + 1, len(detalles) - 1)

    for fila in range(fila_inicio + 1, fila_inicio + len(detalles)):
        copiar_estilo_fila(ws, fila_plantilla, fila)

    for detalle in detalles:

        ws[f"A{fila_contador}"] = contador

        ws[f"B{fila_contador}"] = detalle["badgeNumber"]
        
        ws[f"C{fila_contador}"] = detalle["idEmpleado"]

        ws[f"D{fila_contador}"] =  detalle["nombre_completo"] #f"{detalle["firstName"]} {detalle["secondName"] or ""} {detalle["lastName"]} {detalle["lastName2"] or ""}"

        ws[f"E{fila_contador}"] = detalle["fecha"]

        ws[f"F{fila_contador}"] = detalle["hora_entrada"]

        ws[f"G{fila_contador}"] = detalle["hora_salida"]

        ws[f"H{fila_contador}"] = detalle["hora_entrada_digitada"]

        ws[f"I{fila_contador}"] = detalle["hora_salida_digitada"]
        
        ws[f"J{fila_contador}"] = detalle["hora_entrada_reloj"]
        
        ws[f"K{fila_contador}"] = detalle["hora_entrada_reloj"]
        
        ws[f"L{fila_contador}"] = detalle["total_digitado"]
        
        ws[f"M{fila_contador}"] = detalle["total_horas"]
        
        ws[f"N{fila_contador}"] = detalle["diferencia"]

        autorizado = ""
        if int(detalle["autorizado"]) == 1 and detalle["usuario_autorizacion"] is not None:
            autorizado = "SÍ" 
        elif int(detalle["autorizado"]) == 0 and detalle["usuario_autorizacion"] is not None: 
            autorizado = "NO" 
        else: 
            autorizado = "PENDIENTE"

        ws[f"O{fila_contador}"] = autorizado

        ws[f"P{fila_contador}"] = detalle["horas_autorizadas"] if detalle["horas_autorizadas"] else "----"
        
        ws[f"Q{fila_contador}"] = detalle["nombreUsuario"] if detalle["nombreUsuario"] != "" else "----"
        
        ws[f"R{fila_contador}"] = detalle["fecha_autorizacion"] if detalle["fecha_autorizacion"] else "----"

        fila_contador += 1
        contador += 1

    ws[f"P{fila_contador}"] = f'=COUNT(P{fila_inicio}:P{fila_contador - 1})'

    ws["D5"] = f"=P{fila_contador}"
    ws["D6"] = f'=SUMIFS(N{fila_inicio}:N{fila_contador - 1}, O{fila_inicio}:O{fila_contador - 1}, "PENDIENTE")'

    # =========================================================
    # ANCHO COLUMNAS
    # =========================================================

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 12
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["R"].width = 15

    # =========================================================
    # PIE DE REPORTE
    # =========================================================

    fila_footer = fila_contador + 2

    ws[f"N{fila_footer}"] = "Generado el:"

    ahora = datetime.now(pytz.timezone("America/Guatemala"))

    ws[f"O{fila_footer}"] = ahora.strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws[f"N{fila_footer + 1}"] = "Por:"

    ws[f"O{fila_footer + 1}"] = current_user.nombre

    
    
    # =========================================================
    # EXPORTAR EN MEMORIA
    # =========================================================
    
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return archivo

@login_required
def generar_reporte_horas_extra_pendientes(encabezado_detalles, detalles):
    ruta_plantilla = "app/templates/excel/horas_pendientes.xlsm"

    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws = wb["Pendientes"]

    ws["D2"] = encabezado_detalles["nombreDepartamento"]
    ws["D3"] = encabezado_detalles["from_date"]
    ws["D4"] = encabezado_detalles["to_date"]    
    
    # =========================================================
    # DETALLE EMPLEADOS
    # =========================================================

    fila_inicio = 8
    fila_contador = fila_inicio
    fila_plantilla = 8

    contador = 1

    if len(detalles) > 1:
        ws.insert_rows(fila_inicio + 1, len(detalles) - 1)

    for fila in range(fila_inicio + 1, fila_inicio + len(detalles)):
        copiar_estilo_fila(ws, fila_plantilla, fila)

    for detalle in detalles:

        ws[f"A{fila_contador}"] = contador

        ws[f"B{fila_contador}"] = detalle["idEmpleado"]
        
        ws[f"C{fila_contador}"] = detalle["badgeNumber"]

        ws[f"D{fila_contador}"] =  detalle["nombre_completo"] 

        ws[f"I{fila_contador}"] = detalle["fecha"]

        ws[f"J{fila_contador}"] = detalle["diferencia"]

        fila_contador += 1
        contador += 1

    ws[f"J{fila_contador}"] = f'=SUM(J{fila_inicio}:J{fila_contador - 1})'

    ws["D5"] = f"=J{fila_contador}"

    # =========================================================
    # ANCHO COLUMNAS
    # =========================================================

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 21
    ws.column_dimensions["J"].width = 21
    
    # =========================================================
    # PIE DE REPORTE
    # =========================================================

    fila_footer = fila_contador + 2

    ws[f"I{fila_footer}"] = "Generado el:"

    ahora = datetime.now(pytz.timezone("America/Guatemala"))

    ws[f"J{fila_footer}"] = ahora.strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws[f"I{fila_footer + 1}"] = "Por:"

    ws[f"J{fila_footer + 1}"] = current_user.nombre
    
    # =========================================================
    # EXPORTAR EN MEMORIA
    # =========================================================
    
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return archivo

@login_required
def generar_reporte_resumen_horas_autorizadas(encabezado_detalles, detalles):
    ruta_plantilla = "app/templates/excel/resumen_horas_autorizadas.xlsm"

    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws = wb["Resumen"]

    ws["C2"] = encabezado_detalles["nombreDepartamento"]
    ws["C3"] = encabezado_detalles["from_date"]
    ws["C4"] = encabezado_detalles["to_date"]    
    
    # =========================================================
    # DETALLE EMPLEADOS
    # =========================================================

    fila_inicio = 8
    fila_contador = fila_inicio
    fila_plantilla = 8

    contador = 1

    if len(detalles) > 1:
        ws.insert_rows(fila_inicio + 1, len(detalles) - 1)

    for fila in range(fila_inicio + 1, fila_inicio + len(detalles)):
        copiar_estilo_fila(ws, fila_plantilla, fila)

    for detalle in detalles:

        ws[f"A{fila_contador}"] = contador

        ws[f"B{fila_contador}"] = detalle["idEmpleado"]
        
        ws[f"C{fila_contador}"] = detalle["nombre_completo"]

        ws[f"D{fila_contador}"] =  detalle["suma_horas_autorizadas"] 

        fila_contador += 1
        contador += 1

    ws[f"D{fila_contador}"] = f'=SUM(D{fila_inicio}:D{fila_contador - 1})'

    ws["C5"] = f"=D{fila_contador}"

    # =========================================================
    # ANCHO COLUMNAS
    # =========================================================

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 27

    # =========================================================
    # PIE DE REPORTE
    # =========================================================

    fila_footer = fila_contador + 2

    ws[f"C{fila_footer}"] = "Generado el:"

    ahora = datetime.now(pytz.timezone("America/Guatemala"))

    ws[f"D{fila_footer}"] = ahora.strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws[f"C{fila_footer + 1}"] = "Por:"

    ws[f"D{fila_footer + 1}"] = current_user.nombre
    
    # =========================================================
    # EXPORTAR EN MEMORIA
    # =========================================================
    
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return archivo

@login_required
def generar_reporte_horas_autorizadas_por_empleado_linea(encabezado_detalles, detalles):
    ruta_plantilla = "app/templates/excel/horas_autorizadas_empleado_linea.xlsm"
    fechas = obtener_diferencia_fechas(encabezado_detalles["from_date"], encabezado_detalles["to_date"])

    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws = wb["Resumen_EmpleadoXlinea"]

    ws["C2"] = encabezado_detalles["nombreDepartamento"]
    ws["C3"] = encabezado_detalles["from_date"]
    ws["C4"] = encabezado_detalles["to_date"]    
    
    # =========================================================
    # DETALLE EMPLEADOS
    # =========================================================

    fila_inicio = 8
    fila_contador = fila_inicio
    fila_plantilla = 8

    columna_inicio = 5
    columna_contador = columna_inicio
    columna_plantilla = 5

    contador = 1

    if len(detalles) > 1:
        ws.insert_rows(fila_inicio + 1, len(detalles) - 1)

    for fila in range(fila_inicio + 1, fila_inicio + len(detalles)):
        copiar_estilo_fila(ws, fila_plantilla, fila)


    for fecha in fechas:
        ws.cell(row=fila_inicio - 2, column=columna_contador).value = fecha.strftime("%d/%m/%y")
        ws.cell(row=fila_inicio - 1, column=columna_contador).value = "Horas autorizadas"
        ws.cell(row=fila_inicio - 1, column=columna_contador + 1).value = "Línea asignada"

        columna_contador += 2

    columna_final = columna_contador
    columna_contador = columna_inicio

    ws.cell(row=fila_inicio - 1, column=columna_final).value = "Total horas"

    for detalle in detalles:
        ws[f"A{fila_contador}"] = contador

        ws[f"B{fila_contador}"] = detalle["idEmpleado"]
        
        ws[f"C{fila_contador}"] = detalle["nombre_completo"]

        ws[f"D{fila_contador}"] =  detalle["centro_de_costo"]
        
        horas_total = 0
        for fecha in fechas:
            fecha_key = fecha.strftime("%Y-%m-%d")

            info = detalle["fechas"].get(fecha_key)


            if info:
                ws.cell(row=fila_contador, column=columna_contador).value = info["horas"] if info["horas"] else "----"
                ws.cell(row=fila_contador, column=columna_contador + 1).value = info["linea"] if info["linea"] else "----"
                horas_total += float(info["horas"])
            else:
                ws.cell(row=fila_contador, column=columna_contador).value = "----"
                ws.cell(row=fila_contador, column=columna_contador + 1).value = "----"

            columna_contador += 2

        ws.cell(row=fila_contador, column=columna_final).value = horas_total
        
        columna_contador = columna_inicio

        fila_contador += 1
        contador += 1

    for fecha in fechas:
        column_letter = get_column_letter(columna_contador)
        ws.cell(row=fila_contador, column=columna_contador).value = f"=SUM({column_letter}{fila_inicio}:{column_letter}{fila_contador - 1})"

        columna_contador +=2

    columna_contador = columna_inicio

    final_column_letter = get_column_letter(columna_final)

    ws[f"{final_column_letter}{fila_contador}"] = f"=SUM({final_column_letter}{fila_inicio}:{final_column_letter}{fila_contador - 1})"
    ws["C5"] = f"={final_column_letter}{fila_contador}"

    for i, fecha in enumerate(fechas):
        if i < len(fechas) - 1:
            copiar_estilo_columna(
                ws, 
                columna_plantilla, 
                columna_contador + 2, 
                fila_inicio - 2, 
                fila_contador
            )
            copiar_estilo_columna(
                ws, 
                columna_plantilla + 1, 
                columna_contador + 3, 
                fila_inicio - 2, 
                fila_contador
            )

        columna_contador += 2
    
    copiar_estilo_columna(
        ws, 
        columna_plantilla, 
        columna_final, 
        fila_inicio - 1, 
        fila_contador
    )

    # =========================================================
    # ANCHO COLUMNAS
    # =========================================================

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 30

    # =========================================================
    # PIE DE REPORTE
    # =========================================================

    fila_footer = fila_contador + 2

    ws[f"B{fila_footer}"] = "Generado el:"

    ahora = datetime.now(pytz.timezone("America/Guatemala"))

    ws[f"B{fila_footer}"] = ahora.strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws[f"C{fila_footer + 1}"] = "Por:"

    ws[f"C{fila_footer + 1}"] = current_user.nombre
    
    # =========================================================
    # EXPORTAR EN MEMORIA
    # =========================================================
    
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return archivo

