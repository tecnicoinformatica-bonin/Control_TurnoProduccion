from io import BytesIO
from datetime import datetime, timedelta

from flask_login import login_required, current_user
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

from app.api.centro_de_costo.centro_de_costo_service import Centro_de_costo_Service
from app.extensions.db import db

from copy import copy

import pytz

from app.api.linea.linea_service import Linea_Service

def copiar_estilo_fila(ws, fila_origen, fila_destino):

    for col in range(1, ws.max_column + 1):

        celda_origen = ws.cell(row=fila_origen, column=col)
        celda_destino = ws.cell(row=fila_destino, column=col)

        # Copiar estilo
        celda_destino.font = copy(celda_origen.font)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)

        # Copiar altura fila
        ws.row_dimensions[fila_destino].height = ws.row_dimensions[fila_origen].height

def copiar_estilo_columna(ws, col_origen, col_destino, fila_inicio, fila_fin):

    for fila in range(fila_inicio, fila_fin + 1):

        origen = ws.cell(fila, col_origen)
        destino = ws.cell(fila, col_destino)

        destino.font = copy(origen.font)
        destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format
        destino.protection = copy(origen.protection)

def obtener_diferencia_fechas(from_date, to_date):
        fecha_inicio = datetime.strptime(from_date, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(to_date, "%Y-%m-%d").date()

        fechas = []

        dias = (fecha_fin - fecha_inicio).days + 1

        fechas = [
            fecha_inicio + timedelta(days=i)
            for i in range(dias)
        ]
        
        return fechas

@login_required
def generar_reporte_resumen_motivos_desasignacion(encabezado_detalles, detalles, user, time):
    ruta_plantilla = "app/templates/excel/registro_motivo_desasignacion.xlsm"
    
    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws = wb["Resumen"]  
    
    ws["C3"] = encabezado_detalles["from_date"]
    ws["C4"] = encabezado_detalles["to_date"]    
    
    # =========================================================
    # PARTE 1
    # =========================================================

    fila_inicio = 7
    fila_contador = fila_plantilla = fila_inicio
    
    contador = 1

    if len(detalles) > 1:
        ws.insert_rows(fila_inicio + 1, len(detalles) - 1)

    for fila in range(fila_inicio + 1, fila_inicio + len(detalles)):
        copiar_estilo_fila(ws, fila_plantilla, fila)

    for detalle in detalles:

        ws[f"A{fila_contador}"] = detalle["descripcion"]
        
        ws[f"B{fila_contador}"] = detalle["conteo"]
        
        fila_contador += 1
        contador += 1

    ws[f"C{fila_contador}"] = f"=SUM(B{fila_inicio}:B{fila_contador - 1})"

    # Gráfica
    fila_final = fila_contador - 1
    chart = BarChart()

    chart.type = "col"
    chart.title = f"Motivos por los que RRHH realizó modificaciones"
    chart.style = 10

    # Columna C = Horas Centro
    data = Reference(
        ws,
        min_col=2,      # B
        min_row=fila_inicio,
        max_row=fila_final
    )

    # Columna B = Nombre Centro
    cats = Reference(
        ws,
        min_col=1,      # B
        min_row=fila_inicio,
        max_row=fila_final
    )

    chart.add_data(data)
    chart.set_categories(cats)

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showPercent = False
    chart.dLbls.showLegendKey = False

    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"

    chart.width = 20
    chart.height = 10

    ws.add_chart(chart, "D6")

    # =========================================================
    # PIE DE REPORTE
    # =========================================================

    fila_footer = fila_contador + 2

    ws[f"A{fila_footer}"] = "Generado el:"
    # ws[f"B{fila_footer}"] = "Generado el:"

    ws[f"B{fila_footer}"] = time.strftime(
    # ws[f"B{fila_footer}"] = ahora.strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws[f"A{fila_footer + 1}"] = "Por:"
    # ws[f"C{fila_footer + 1}"] = "Por:"

    ws[f"B{fila_footer + 1}"] = user
    # ws[f"C{fila_footer + 1}"] = current_user.nombre
    
    # =========================================================
    # EXPORTAR EN MEMORIA
    # =========================================================
    
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return archivo

